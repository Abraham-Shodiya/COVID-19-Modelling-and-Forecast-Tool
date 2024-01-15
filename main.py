import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
import seaborn as sns

wb = load_workbook("ExamResults.xlsx")
ws = wb.active
ws.delete_rows(1, amount=2)
ws.delete_cols(1, amount=1)


def get_points(student, course):
    global cell
    tot = 0
    count = 0

    for i in range(1, 21):

        try:
            cell = int(ws.cell(student, i).value)
            # print(cell)
            tot += cell
            count += 1

        except ValueError:
            continue

        if i == course:
            tot -= cell
            count -= 1

    return round(tot / count), round(int(ws.cell(student, course).value))


x1 = []
y1 = []

x2 = []
y2 = []
for i in range(2, 101):
    try:
        print(get_points(i, 1))
        x = get_points(i, 1)[0]
        y = get_points(i, 1)[1]
        xx = get_points(i, 2)[0]
        yy = get_points(i, 2)[1]

        x1.append(x)
        x2.append(xx)
        y1.append(y)
        y2.append(yy)

    except ValueError:
        continue


plot1 = plt.figure(1)
sns.regplot(x1, y2, ci=None)

plot2 = plt.figure(2)
sns.regplot(x2, y2, ci=None)

plt.show()